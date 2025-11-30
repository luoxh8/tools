import json

import openpyxl

# 加载 json 数据
with open('data.json', 'r') as f:
    data = json.load(f)

# 创建一个新的工作簿
wb = openpyxl.Workbook()

# 删除默认的工作表
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# 遍历 sheet_name 创建对应的 sheet
for sheet_name in data['sheet_name']:
    wb.create_sheet(sheet_name)

# 创建一个名为 'others' 的 sheet
wb.create_sheet('others')

# 遍历 data 中的 obj
for obj in data['data']:
    # 获取 belong 属性
    belong = obj['belong']
    # 获取 CRNumber 属性
    cr_number = obj['CRNumber']
    # 检查 belong 是否匹配 sheet_name 中的任意一个
    matched = False
    for sheet_name in data['sheet_name']:
        if belong.startswith(sheet_name):
            # 如果匹配，则将 obj 放在对应的 sheet 中
            ws = wb[sheet_name]
            ws.append([belong, cr_number])
            matched = True
            break
    if not matched:
        # 如果不匹配，则将 obj 放在 'others' 中
        ws = wb['others']
        ws.append([belong, cr_number])

# 保存工作簿
wb.save('output.xlsx')
